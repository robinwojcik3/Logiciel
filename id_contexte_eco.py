import os
import sys
import datetime
import math
import geopandas as gpd
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def run_id_contexte_eco(couche_reference1: str, couche_reference2: str) -> None:
    """Exécute l'identification des zonages et exporte un Excel.

    :param couche_reference1: chemin vers l'aire d'étude élargie
    :param couche_reference2: chemin vers la zone d'étude
    """
    qgis_base = r"C:\Program Files\QGIS 3.40.3"
    os.environ['PROJ_LIB'] = os.path.join(qgis_base, "share", "proj")
    os.environ['PROJ_DATA'] = os.path.join(qgis_base, "share", "proj")
    os.environ['GDAL_DATA'] = os.path.join(qgis_base, "share", "gdal")

    def log_with_time(message: str) -> None:
        now = datetime.datetime.now().strftime("%H:%M:%S")
        print(f"[{now}] {message}")

    ascii_art = ""

    log_with_time("Démarrage du script d'identification des zonages...")

    if not os.path.exists(couche_reference1):
        log_with_time(f"Le fichier de la première couche de référence n'a pas été trouvé : {couche_reference1}")
        return

    if not os.path.exists(couche_reference2):
        log_with_time(f"Le fichier de la deuxième couche de référence n'a pas été trouvé : {couche_reference2}")
        return

    try:
        reference_gdf = gpd.read_file(couche_reference1)
        log_with_time("Première couche de référence chargée avec succès")
    except Exception as e:
        log_with_time(f"Erreur lors du chargement de la première couche de référence : {e}")
        return

    try:
        reference2_gdf = gpd.read_file(couche_reference2)
        log_with_time("Deuxième couche de référence chargée avec succès")
    except Exception as e:
        log_with_time(f"Erreur lors du chargement de la deuxième couche de référence : {e}")
        return

    crs_projected = "EPSG:2154"

    if reference_gdf.crs != crs_projected:
        try:
            reference_gdf = reference_gdf.to_crs(crs_projected)
            log_with_time("Reprojection de la première couche de référence effectuée")
        except Exception as e:
            log_with_time(f"Erreur lors de la reprojection de la première couche de référence : {e}")
            return

    if reference2_gdf.crs != crs_projected:
        try:
            reference2_gdf = reference2_gdf.to_crs(crs_projected)
            log_with_time("Reprojection de la deuxième couche de référence effectuée")
        except Exception as e:
            log_with_time(f"Erreur lors de la reprojection de la deuxième couche de référence : {e}")
            return

    try:
        reference2_centroid = reference2_gdf.geometry.union_all().centroid
        log_with_time("Calcul du centroïde de référence effectué")
    except Exception as e:
        log_with_time(f"Erreur lors du calcul du centroïde de la deuxième couche de référence : {e}")
        return

    def calculate_azimuth(point1, point2):
        delta_x = point2.x - point1.x
        delta_y = point2.y - point1.y
        angle_rad = math.atan2(delta_x, delta_y)
        angle_deg = math.degrees(angle_rad)
        azimuth = (angle_deg + 360) % 360
        return azimuth

    def map_azimuth_to_direction(azimuth):
        if (337.5 <= azimuth < 360) or (0 <= azimuth < 22.5):
            return 'Nord'
        elif 22.5 <= azimuth < 67.5:
            return 'Nord-est'
        elif 67.5 <= azimuth < 112.5:
            return 'Est'
        elif 112.5 <= azimuth < 157.5:
            return 'Sud-est'
        elif 157.5 <= azimuth < 202.5:
            return 'Sud'
        elif 202.5 <= azimuth < 247.5:
            return 'Sud-ouest'
        elif 247.5 <= azimuth < 292.5:
            return 'Ouest'
        elif 292.5 <= azimuth < 337.5:
            return 'Nord-ouest'
        else:
            return 'Inconnu'

    def combine_distance_and_direction(distance_km, direction):
        if distance_km == 0.0:
            return "Se superpose à la zone d'étude"
        if direction in ['Est', 'Ouest']:
            preposition = "à l'"
            direction_str = direction.lower()
        else:
            preposition = 'au '
            direction_str = direction.lower()
        combined_str = f"{distance_km} km {preposition}{direction_str}"
        return combined_str

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', start_color='4F81BD', end_color='4F81BD')
    data_font = Font(name='Calibri', size=11)
    data_fill_even = PatternFill(fill_type='solid', start_color='F2F2F2', end_color='F2F2F2')
    data_fill_odd = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    couches_cibles = [
        {
            'nom': 'N2000 ZPS',
            'chemin': r"C:\Users\utilisateur\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\N2000 ZPS.shp",
            'attributs': ['SITENAME', 'SITECODE'],
        },
        {
            'nom': 'N2000 ZSC',
            'chemin': r"C:\Users\utilisateur\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\N2000 ZSC.shp",
            'attributs': ['SITENAME', 'SITECODE'],
        },
        {
            'nom': 'ZNIEFF de Type I',
            'chemin': r"C:\Users\utilisateur\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZNIEFF de type I.shp",
            'attributs': ['NOM','ID_MNHN','ID_ORG'],
        },
        {
            'nom': 'ZNIEFF de Type II',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZNIEFF de type II.shp",
            'attributs': ['NOM','ID_MNHN','ID_ORG'],
        },
        {
            'nom': 'APPB',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\APPB.shp",
            'attributs': ['NOM_SITE','ID_MNHN','URL_FICHE','OPERATEUR'],
        },
        {
            'nom': 'APPHN',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\APPHN.shp",
            'attributs': ['NOM_SITE','ID_MNHN','URL_FICHE','OPERATEUR'],
        },
        {
            'nom': 'Terrain CEN - Terrain gere',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Terrain CEN - Terrain gere.shp",
            'attributs': ['ID_MNHN','NOM_SITE'],
        },
        {
            'nom': 'Terrain CEN - Terrain acquis',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Terrain CEN - Terrain acquis.shp",
            'attributs': ['ID_MNHN','NOM_SITE'],
        },
        {
            'nom': 'ENS',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ENS.shp",
            'attributs': ['NOM_SITE','ID_MNHN','URL_FICHE','GEST_SITE','OPERATEUR','STAT_FON'],
        },
        {
            'nom': 'Parc Nationaux',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Parc Nationaux.shp",
            'attributs': ['NOM_SITE','ID_MNHN','ID_LOCAL','PPN_ASSO','URL_FICHE','GEST_SITE','OPERATEUR'],
        },
        {
            'nom': 'Parc Naturels Régionaux',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Parc Naturels Regionaux.shp",
            'attributs': ['NOM_SITE','ID_MNHN','GEST_SITE','URL_FICHE'],
        },
        {
            'nom': 'Réserve biologique',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Réserve biologique.shp",
            'attributs': ['NOM_SITE','ID_MNHN','GEST_SITE','URL_FICHE'],
        },
        {
            'nom': 'Réserve de biosphère',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Réserve de biosphère.shp",
            'attributs': ['NOM_SITE','ID_MNHN','GEST_SITE','URL_FICHE','OPERATEUR'],
        },
        {
            'nom': 'Réserve intégrale de PN',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Réserve intégrale de PN.shp",
            'attributs': ['NOM_SITE','ID_MNHN','GEST_SITE','URL_FICHE','OPERATEUR','ID_PN'],
        },
        {
            'nom': 'Réserve nationale',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Réserve nationale.shp",
            'attributs': ['NOM_SITE','ID_MNHN','ID_LOCAL','URL_FICHE','ACTE_DEB','GEST_SITE','OPERATEUR'],
        },
        {
            'nom': 'Réserve régionale',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Réserve régionale.shp",
            'attributs': ['NOM_SITE','ID_MNHN','ID_LOCAL','URL_FICHE','ACTE_DEB','GEST_SITE','OPERATEUR'],
        },
        {
            'nom': 'ZH 01',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZH 01.shp",
            'attributs': ['nom','id_map','id_local','url'],
        },
        {
            'nom': 'ZH 26',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZH 26.shp",
            'attributs': ['site_name','nom_bv','site_cod','sdage'],
        },
        {
            'nom': 'ZH 38',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZH 38.shp",
            'attributs': ['nom','id_map','id_local','url'],
        },
        {
            'nom': 'ZH 69',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZH 69.shp",
            'attributs': ['nom','id_map','id_local','url'],
        },
        {
            'nom': 'ZH 73',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZH 73.shp",
            'attributs': ['site_name','id_bdd'],
        },
        {
            'nom': 'ZH 74',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZH 74.shp",
            'attributs': ['NOM'],
        },
        {
            'nom': 'ZH Bourgogne',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZH Bourgogne.shp",
            'attributs': ['nom','id_map','id_local','url'],
        },
        {
            'nom': 'ZH PACA',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\ZH PACA.shp",
            'attributs': ['site','code','lib_ssbv','type_sdage'],
        },
        {
            'nom': 'Pelouses sèches 38',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Pelouses sèches 38.dbf",
            'attributs': ['LEGENDE','ID'],
        },
        {
            'nom': 'Pelouses sèches 73',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Pelouses sèches 73.dbf",
            'attributs': ['site_name','id_bdd'],
        },
        {
            'nom': 'Pelouses sèches 74',
            'chemin': r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\INPUT\Tables pour ID zonages\Pelouses sèches 74.dbf",
            'attributs': ['Site','ID'],
        }
    ]

    dossier_sortie = r"C:\USERS\UTILISATEUR\Mon Drive\1 - Bota & Travail\+++++++++  BOTA  +++++++++\---------------------- 3) BDD\PYTHON\2) Contexte éco\OUTPUT"
    nom_fichier_sortie = 'ID zonages.xlsx'
    chemin_sortie = os.path.join(dossier_sortie, nom_fichier_sortie)

    if os.path.exists(chemin_sortie):
        try:
            os.remove(chemin_sortie)
            log_with_time(f"Le fichier existant '{chemin_sortie}' a été supprimé et sera remplacé par le nouvel export.")
        except Exception as e:
            log_with_time(f"Erreur lors de la suppression du fichier existant '{chemin_sortie}': {e}")
            return

    def process_single_layer(couche, writer):
        sheet_name = couche['nom']
        nom_couche = couche['nom']
        chemin_cible = couche['chemin']
        attributs_a_exporter = couche['attributs']
        row_position = 0
        if not os.path.exists(chemin_cible):
            log_with_time(f"Le fichier de la couche '{nom_couche}' n'a pas été trouvé : {chemin_cible}")
            return
        try:
            cible_gdf = gpd.read_file(chemin_cible)
            log_with_time(f"Couche '{nom_couche}' chargée")
        except Exception as e:
            log_with_time(f"Erreur lors du chargement de la couche '{nom_couche}': {e}")
            return
        if cible_gdf.crs != crs_projected:
            try:
                cible_gdf = cible_gdf.to_crs(crs_projected)
                log_with_time(f"Reprojection de '{nom_couche}' effectuée")
            except Exception as e:
                log_with_time(f"Erreur lors de la reprojection de '{nom_couche}': {e}")
                return
        try:
            overlapping_gdf = gpd.sjoin(cible_gdf, reference_gdf, how='inner', predicate='intersects')
            log_with_time(f"Jointure spatiale pour '{nom_couche}' effectuée")
        except Exception as e:
            log_with_time(f"Erreur lors de la jointure spatiale pour '{nom_couche}': {e}")
            return
        if overlapping_gdf.empty:
            log_with_time(f"Aucun site présent dans '{nom_couche}'.")
            return
        try:
            distances = overlapping_gdf.geometry.apply(lambda geom: reference2_gdf.distance(geom).min())
            distances_km = distances / 1000
            distances_km = distances_km.round(1)
            overlapping_gdf['Distance (km)'] = distances_km
            log_with_time(f"Calcul des distances pour '{nom_couche}' effectué")
        except Exception as e:
            log_with_time(f"Erreur lors du calcul des distances pour '{nom_couche}': {e}")
            return
        try:
            overlapping_gdf['centroid'] = overlapping_gdf.geometry.centroid
            overlapping_gdf['Azimuth (°)'] = overlapping_gdf['centroid'].apply(lambda geom: calculate_azimuth(reference2_centroid, geom))
            overlapping_gdf['Azimuth'] = overlapping_gdf['Azimuth (°)'].apply(map_azimuth_to_direction)
            log_with_time(f"Calcul de l'azimut pour '{nom_couche}' effectué")
        except Exception as e:
            log_with_time(f"Erreur lors du calcul de l'azimut pour '{nom_couche}': {e}")
            return
        try:
            overlapping_gdf['Distance et Direction'] = overlapping_gdf.apply(
                lambda row: combine_distance_and_direction(row['Distance (km)'], row['Azimuth']),
                axis=1
            )
            log_with_time(f"Combinaison distance/direction pour '{nom_couche}' effectuée")
        except Exception as e:
            log_with_time(f"Erreur lors de la combinaison de la distance et de la direction pour '{nom_couche}': {e}")
            return
        available_columns = overlapping_gdf.columns
        attributs_existants = []
        colonne_mapping = {}
        for attr in attributs_a_exporter:
            found = False
            for col in available_columns:
                if col.lower() == attr.lower():
                    attributs_existants.append(col)
                    colonne_mapping[attr] = col
                    found = True
                    break
            if not found:
                log_with_time(f"L'attribut '{attr}' n'a pas été trouvé dans '{nom_couche}'.")
        if not attributs_existants:
            log_with_time(f"Aucun des attributs spécifiés n'a été trouvé dans '{nom_couche}'.")
            log_with_time(f"Attributs disponibles dans '{nom_couche}': {available_columns.tolist()}")
            return
        try:
            resultats = overlapping_gdf[['Distance et Direction'] + attributs_existants].copy()
            log_with_time(f"Extraction des attributs pour '{nom_couche}' effectuée")
        except Exception as e:
            log_with_time(f"Erreur lors de l'extraction des attributs dans '{nom_couche}': {e}")
            return
        for attr, col in colonne_mapping.items():
            resultats.rename(columns={col: attr}, inplace=True)
        resultats['Distance numérique'] = overlapping_gdf['Distance (km)']
        resultats = resultats.sort_values(by='Distance numérique', ascending=True)
        resultats.drop(columns='Distance numérique', inplace=True)
        resultats.insert(0, 'Nom de la couche', nom_couche)
        colonnes_df = ['Nom de la couche', 'Distance et Direction'] + attributs_a_exporter
        for col in colonnes_df:
            if col not in resultats.columns:
                resultats[col] = ''
        resultats = resultats[colonnes_df]
        nombre_total = len(overlapping_gdf)
        ligne_nom_couche = pd.DataFrame(
            [[nom_couche, f"Nombre de {nom_couche} dans l'aire d'étude élargie : {nombre_total}"] + [''] * (len(colonnes_df) - 2)],
            columns=colonnes_df
        )
        ligne_noms_attributs = pd.DataFrame([colonnes_df], columns=colonnes_df)
        start_row = row_position
        ligne_nom_couche.to_excel(writer, sheet_name=sheet_name, startrow=row_position, header=False, index=False)
        row_position += 1
        ligne_noms_attributs.to_excel(writer, sheet_name=sheet_name, startrow=row_position, header=False, index=False)
        header_row_number = row_position
        row_position += 1
        data_start_row = row_position
        resultats.to_excel(writer, sheet_name=sheet_name, startrow=row_position, header=False, index=False)
        row_position += len(resultats)
        data_end_row = row_position - 1
        worksheet = writer.sheets[sheet_name]
        start_col = 1
        end_col = len(colonnes_df)
        for col_idx in range(start_col, end_col + 1):
            cell = worksheet.cell(row=start_row + 1, column=col_idx)
            cell.font = Font(name='Calibri', size=12, bold=True)
            cell.alignment = alignment
        for col_idx in range(start_col, end_col + 1):
            cell = worksheet.cell(row=header_row_number + 1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment
        for row_idx in range(data_start_row + 1, data_end_row + 2):
            for col_idx in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = border
                cell.alignment = alignment
                if (row_idx - data_start_row) % 2 == 0:
                    cell.fill = data_fill_even
                else:
                    cell.fill = data_fill_odd
        for col_idx in range(start_col, end_col + 1):
            column_letter = get_column_letter(col_idx)
            if col_idx == 3:
                worksheet.column_dimensions[column_letter].width = 45
            elif 4 <= col_idx <= 9:
                worksheet.column_dimensions[column_letter].width = 15
            else:
                worksheet.column_dimensions[column_letter].width = 20
        log_with_time(f"Traitement de la couche '{nom_couche}' terminé")

    def process_synthesis(couches, sheet_name, writer):
        log_with_time(f"Création de la feuille de synthèse...")
        all_results = []
        code_field_mapping = {
            'N2000 ZPS': 'SITECODE',
            'N2000 ZSC': 'SITECODE',
            'ZNIEFF de Type I': 'ID_MNHN',
            'ZNIEFF de Type II': 'ID_MNHN'
        }
        for couche in couches:
            nom_couche = couche['nom']
            chemin_cible = couche['chemin']
            if nom_couche in ['ZH 38', 'Pelouses sèches']:
                if not os.path.exists(chemin_cible):
                    log_with_time(f"Le fichier de la couche '{nom_couche}' n'a pas été trouvé : {chemin_cible}")
                    continue
                try:
                    cible_gdf = gpd.read_file(chemin_cible)
                except Exception as e:
                    log_with_time(f"Erreur lors du chargement de la couche '{nom_couche}': {e}")
                    continue
                if cible_gdf.crs != crs_projected:
                    try:
                        cible_gdf = cible_gdf.to_crs(crs_projected)
                    except Exception as e:
                        log_with_time(f"Erreur lors de la reprojection de '{nom_couche}': {e}")
                        continue
                try:
                    overlapping_gdf = gpd.sjoin(cible_gdf, reference_gdf, how='inner', predicate='intersects')
                except Exception as e:
                    log_with_time(f"Erreur lors de la jointure spatiale pour '{nom_couche}': {e}")
                    continue
                if overlapping_gdf.empty:
                    continue
                nombre_objets = len(overlapping_gdf)
                if nom_couche == 'ZH 38':
                    type_zonage = 'Zone humide'
                    nom_du_site = f"{nombre_objets} zones humides dans l'aire d'étude élargie"
                elif nom_couche == 'Pelouses sèches':
                    type_zonage = 'Pelouses sèches'
                    nom_du_site = f"{nombre_objets} pelouses sèches dans l'aire d'étude élargie"
                temp_df = pd.DataFrame({
                    'Type de zonage': [type_zonage],
                    'Distance à la zone d\'étude': ['/'],
                    'Nom du site': [nom_du_site],
                    '': [''],
                    'CODES': ['']
                })
                all_results.append(temp_df)
                continue
            if not os.path.exists(chemin_cible):
                log_with_time(f"Le fichier de la couche '{nom_couche}' n'a pas été trouvé : {chemin_cible}")
                continue
            try:
                cible_gdf = gpd.read_file(chemin_cible)
            except Exception as e:
                log_with_time(f"Erreur lors du chargement de la couche '{nom_couche}': {e}")
                continue
            if cible_gdf.crs != crs_projected:
                try:
                    cible_gdf = cible_gdf.to_crs(crs_projected)
                except Exception as e:
                    log_with_time(f"Erreur lors de la reprojection de '{nom_couche}': {e}")
                    continue
            try:
                overlapping_gdf = gpd.sjoin(cible_gdf, reference_gdf, how='inner', predicate='intersects')
            except Exception as e:
                log_with_time(f"Erreur lors de la jointure spatiale pour '{nom_couche}': {e}")
                continue
            if overlapping_gdf.empty:
                continue
            try:
                distances = overlapping_gdf.geometry.apply(lambda geom: reference2_gdf.distance(geom).min())
                distances_km = distances / 1000
                distances_km = distances_km.round(1)
                overlapping_gdf['Distance (km)'] = distances_km
            except Exception as e:
                log_with_time(f"Erreur lors du calcul des distances pour '{nom_couche}': {e}")
                continue
            try:
                overlapping_gdf['centroid'] = overlapping_gdf.geometry.centroid
                overlapping_gdf['Azimuth (°)'] = overlapping_gdf['centroid'].apply(lambda geom: calculate_azimuth(reference2_centroid, geom))
                overlapping_gdf['Azimuth'] = overlapping_gdf['Azimuth (°)'].apply(map_azimuth_to_direction)
            except Exception as e:
                log_with_time(f"Erreur lors du calcul de l'azimut pour '{nom_couche}': {e}")
                continue
            try:
                overlapping_gdf['Distance et Direction'] = overlapping_gdf.apply(
                    lambda row: combine_distance_and_direction(row['Distance (km)'], row['Azimuth']),
                    axis=1
                )
            except Exception as e:
                log_with_time(f"Erreur lors de la combinaison de la distance et de la direction pour '{nom_couche}': {e}")
                continue
            if nom_couche in ['ZNIEFF de Type I', 'ZNIEFF de Type II']:
                name_attr = 'NOM'
            elif nom_couche == 'ZH PACA':
                name_attr = 'site'
            elif nom_couche == 'Pelouses sèches 73':
                name_attr = 'site_name'
            elif nom_couche == 'Pelouses sèches 74':
                name_attr = 'Site'
            elif nom_couche == 'ZH 73':
                name_attr = 'site_name'
            elif nom_couche == 'ZH 74':
                name_attr = 'NOM'
            else:
                name_attr = None
                for attr in ['SITENAME', 'NAME', 'NOM_SITE', 'nom', 'site']:
                    if attr in overlapping_gdf.columns:
                        name_attr = attr
                        break
            if not name_attr:
                log_with_time(f"L'attribut 'Nom du site' est manquant pour '{nom_couche}'.")
                continue
            if nom_couche in code_field_mapping:
                code_field = code_field_mapping[nom_couche]
                if code_field in overlapping_gdf.columns:
                    overlapping_gdf['CODES'] = overlapping_gdf[code_field].astype(str)
                else:
                    overlapping_gdf['CODES'] = ''
                    log_with_time(f"L'attribut '{code_field}' n'a pas été trouvé dans '{nom_couche}'.")
            else:
                overlapping_gdf['CODES'] = ''
            try:
                temp_df = overlapping_gdf[['Distance et Direction', name_attr, 'CODES']].copy()
            except Exception as e:
                log_with_time(f"Erreur lors de l'extraction des attributs dans '{nom_couche}': {e}")
                continue
            temp_df.rename(columns={
                'Distance et Direction': "Distance à la zone d'étude",
                name_attr: 'Nom du site'
            }, inplace=True)
            temp_df.insert(0, 'Type de zonage', nom_couche)
            temp_df.insert(3, '', '')
            all_results.append(temp_df)
        if all_results:
            results_df = pd.concat(all_results, ignore_index=True)
            results_df['Type de zonage'] = results_df['Type de zonage'].replace({'ZH 38': 'Zone humide'})
            type_zonage_order = [couche['nom'] for couche in couches]
            type_zonage_order = ['Zone humide' if nom == 'ZH 38' else nom for nom in type_zonage_order]
            results_df['Type de zonage'] = pd.Categorical(results_df['Type de zonage'], categories=type_zonage_order, ordered=True)
            results_df['Distance numérique'] = results_df["Distance à la zone d'étude"].apply(
                lambda x: float(x.split(' ')[0]) if x != '/' and not x.startswith("Se") else float('inf')
            )
            results_df.sort_values(by=['Type de zonage', 'Distance numérique'], ascending=[True, True], inplace=True)
            results_df.drop(columns=['Distance numérique'], inplace=True)
            results_df.to_excel(writer, sheet_name=sheet_name, startrow=1, header=False, index=False)
            worksheet = writer.sheets[sheet_name]
            headers = ['Type de zonage', "Distance à la zone d'étude", 'Nom du site', '', 'CODES']
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment
            data_start_row = 2
            data_end_row = data_start_row + len(results_df) - 1
            start_col = 1
            end_col = 5
            for row_idx in range(data_start_row, data_end_row + 1):
                for col_idx in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = alignment
                    if (row_idx - data_start_row) % 2 == 0:
                        cell.fill = data_fill_even
                    else:
                        cell.fill = data_fill_odd
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 50
            worksheet.column_dimensions['D'].width = 5
            worksheet.column_dimensions['E'].width = 20
            log_with_time(f"Feuille de synthèse créée avec {len(results_df)} zonages")
        else:
            log_with_time(f"Aucun résultat à écrire dans la feuille '{sheet_name}'. Création d'une feuille vide.")
            empty_df = pd.DataFrame(columns=['Type de zonage', "Distance à la zone d'étude", 'Nom du site', '', 'CODES'])
            empty_df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for col_idx, header in enumerate(empty_df.columns, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 50
            worksheet.column_dimensions['D'].width = 5
            worksheet.column_dimensions['E'].width = 20

    try:
        heure_debut = datetime.datetime.now()
        log_with_time(f"Début de la création du fichier Excel: {chemin_sortie}")
        with pd.ExcelWriter(chemin_sortie, engine='openpyxl') as writer:
            process_synthesis(couches_cibles, 'SYNTHÈSE', writer)
            log_with_time(f"Traitement des onglets individuels ({len(couches_cibles)} couches)...")
            for i, couche in enumerate(couches_cibles):
                log_with_time(f"Traitement de la couche {i+1}/{len(couches_cibles)}: {couche['nom']}")
                process_single_layer(couche, writer)
        duree_totale = datetime.datetime.now() - heure_debut
        log_with_time(f"Fichier Excel créé avec succès en {duree_totale}")
    except Exception as e:
        log_with_time(f"Erreur lors de l'écriture dans le fichier Excel : {e}")
        return

    log_with_time(f"\nLes résultats ont été exportés avec succès dans le fichier Excel : {chemin_sortie}")
    print(ascii_art)
